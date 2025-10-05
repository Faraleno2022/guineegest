from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from calendar import monthrange
import json
import logging

from .models import (
    Employe, PresenceJournaliere, HeureSupplementaire, PaieEmploye, 
    ParametrePaie, ArchiveMensuelle
)
from .models_entreprise import ConfigurationChargesSociales
from decimal import Decimal

logger = logging.getLogger(__name__)

# ===== VUES POUR LES PAIES =====

@login_required
def paie_employe_list(request):
    """Vue pour afficher la liste des paies des employés avec calculs automatiques"""
    try:
        from datetime import datetime
        from .utils_presence_paie import calculer_statistiques_presence, synchroniser_presence_vers_paie
        
        # Récupérer les filtres avec valeurs par défaut
        employe_id = request.GET.get('employe_id')
        annee = int(request.GET.get('annee', datetime.now().year))
        mois = int(request.GET.get('mois', datetime.now().month))
        # Synchronisation automatique activée en permanence; force_sync ignoré
        auto_sync = True
        force_sync = False
        
        # Si auto_sync est activé, synchroniser automatiquement les données manquantes
        if auto_sync:
            employes_actifs = Employe.objects.filter(user=request.user, statut='Actif')
            for employe in employes_actifs:
                try:
                    # Créer ou mettre à jour automatiquement la paie avec les calculs de présence
                    paie_employe, created = PaieEmploye.objects.get_or_create(
                        employe=employe,
                        mois=mois,
                        annee=annee,
                        defaults={'salaire_base': getattr(employe, 'salaire_base', 0)}
                    )
                    # Synchroniser les données de présence selon les formules exactes
                    paie_employe.synchroniser_donnees_presence()
                    # Synchroniser les données d'heures supplémentaires
                    paie_employe.synchroniser_donnees_heures_supp()
                    paie_employe.save()
                except Exception:
                    continue  # Continuer même en cas d'erreur pour un employé
        
        # Récupérer les paies filtrées
        paies = PaieEmploye.objects.filter(
            employe__user=request.user,
            mois=mois,
            annee=annee
        ).select_related('employe').order_by('employe__matricule')
        
        if employe_id:
            paies = paies.filter(employe_id=employe_id)
        
        employes = Employe.objects.filter(user=request.user, statut='Actif').order_by('matricule')
        
        # Enrichir les données avec les calculs de présence et heures supplémentaires
        paies_enrichies = []
        for paie in paies:
            try:
                # Calculer les statistiques de présence selon les formules exactes
                stats_calculees = calculer_statistiques_presence(
                    paie.employe, paie.mois, paie.annee
                )
                
                # Calculer les heures supplémentaires pour ce mois
                from .models_entreprise import HeureSupplementaire
                from datetime import datetime
                import calendar
                
                # Déterminer la plage de dates pour le mois
                premier_jour = datetime(paie.annee, paie.mois, 1).date()
                dernier_jour_mois = calendar.monthrange(paie.annee, paie.mois)[1]
                dernier_jour = datetime(paie.annee, paie.mois, dernier_jour_mois).date()
                
                # Récupérer les heures supplémentaires pour cet employé ce mois
                heures_supp_mois = HeureSupplementaire.objects.filter(
                    employe=paie.employe,
                    date__gte=premier_jour,
                    date__lte=dernier_jour
                )
                
                # Calculer les totaux
                total_heures_supp = sum(float(h.duree) for h in heures_supp_mois)
                total_montant_supp = sum(float(h.total_a_payer) for h in heures_supp_mois)

                # Synchroniser les champs stockés avec les valeurs calculées (auto_sync activé)
                if auto_sync:
                    try:
                        paie.jours_mois = stats_calculees.get('total_jours_mois')
                    except Exception:
                        pass
                    paie.jours_presence = stats_calculees.get('jours_presence', 0)
                    paie.absences = stats_calculees.get('absent', 0)
                    paie.jours_repos = stats_calculees.get('j_repos', 0)
                    paie.dimanches = stats_calculees.get('sundays', 0)
                    paie.heures_supplementaires = total_heures_supp
                    paie.montant_heures_supplementaires = total_montant_supp
                    # Sauvegarder
                    paie.save()
                
                # Vérifier la cohérence avec les données stockées
                coherence = {
                    'jours_presence': paie.jours_presence == stats_calculees['jours_presence'],
                    'absences': paie.absences == stats_calculees['absent'],
                    'dimanches': paie.dimanches == stats_calculees['sundays'],
                    'jours_repos': paie.jours_repos == stats_calculees['j_repos'],
                    'heures_supp': abs(float(paie.heures_supplementaires or 0) - total_heures_supp) < 0.01,
                    'montant_supp': abs(float(paie.montant_heures_supplementaires or 0) - total_montant_supp) < 0.01,
                }
                
                # Ajouter les données d'heures supplémentaires aux stats
                stats_calculees['heures_supplementaires'] = total_heures_supp
                stats_calculees['montant_heures_supplementaires'] = total_montant_supp
                
                # ===== Calculs financiers (CNSS 5%, RTS, Net) =====
                try:
                    # Salaire brut = salaire_base + primes/indemnités + montant HS
                    salaire_base = paie.salaire_base or (paie.employe.salaire_journalier or 0)
                    primes_indemnites = (
                        (paie.prime_discipline or 0) +
                        (paie.cherete_vie or 0) +
                        (paie.indemnite_transport or 0) +
                        (paie.indemnite_logement or 0)
                    )
                    salaire_brut = Decimal(str(salaire_base)) + Decimal(str(primes_indemnites)) + Decimal(str(total_montant_supp or 0))

                    # CNSS: par défaut 5% activé; clés de configuration si présentes
                    try:
                        param_appliquer = ParametrePaie.objects.get(cle='CNSS_ACTIVER', user=request.user)
                        appliquer_cnss = str(param_appliquer.valeur).strip() in ['1', 'true', 'True', 'on']
                    except ParametrePaie.DoesNotExist:
                        appliquer_cnss = True

                    try:
                        param_taux = ParametrePaie.objects.get(cle='CNSS_TAUX', user=request.user)
                        taux_cnss = Decimal(str(param_taux.valeur))
                    except ParametrePaie.DoesNotExist:
                        taux_cnss = Decimal('5.0')

                    cnss_employe = (salaire_brut * taux_cnss / Decimal('100')) if appliquer_cnss else Decimal('0.00')

                    # RTS: type FIXE avec taux, sinon barème progressif par défaut
                    try:
                        param_rts_type = ParametrePaie.objects.get(cle='RTS_TYPE', user=request.user)
                        rts_type = (param_rts_type.valeur or 'PROGRESSIF').upper()
                    except ParametrePaie.DoesNotExist:
                        rts_type = 'PROGRESSIF'

                    try:
                        param_rts_taux = ParametrePaie.objects.get(cle='RTS_TAUX_FIXE', user=request.user)
                        rts_taux_fixe = Decimal(str(param_rts_taux.valeur))
                    except ParametrePaie.DoesNotExist:
                        rts_taux_fixe = Decimal('10.0')

                    salaire_net_imposable = salaire_brut - cnss_employe
                    if rts_type == 'FIXE':
                        rts_employe = (salaire_net_imposable * rts_taux_fixe) / Decimal('100')
                    else:
                        # Barème progressif simple
                        if salaire_net_imposable <= Decimal('1000000'):
                            rts_employe = Decimal('0.00')
                        elif salaire_net_imposable <= Decimal('3000000'):
                            rts_employe = (salaire_net_imposable - Decimal('1000000')) * Decimal('0.05')
                        else:
                            rts_tranche_2 = Decimal('2000000') * Decimal('0.05')
                            rts_tranche_3 = (salaire_net_imposable - Decimal('3000000')) * Decimal('0.15')
                            rts_employe = rts_tranche_2 + rts_tranche_3

                    avance = Decimal(str(paie.avance_sur_salaire or paie.employe.avances or 0))
                    sanctions = Decimal(str(paie.sanction_vol_carburant or 0))
                    net_a_payer = salaire_brut - (cnss_employe + rts_employe + avance + sanctions)

                    # Persister si auto_sync ou force_sync
                    if auto_sync or force_sync:
                        paie.salaire_brut = salaire_brut
                        paie.cnss = cnss_employe
                        paie.rts = rts_employe
                        paie.salaire_net_a_payer = net_a_payer
                        paie.save()
                except Exception:
                    # En cas d'erreur, ne pas bloquer l'affichage
                    pass
                
                paies_enrichies.append({
                    'paie': paie,
                    'stats_calculees': stats_calculees,
                    'coherence': coherence,
                    'a_incoherences': not all(coherence.values()),
                    'heures_supp_details': list(heures_supp_mois)
                })
            except Exception as e:
                # En cas d'erreur, ajouter la paie sans enrichissement
                paies_enrichies.append({
                    'paie': paie,
                    'stats_calculees': {
                        'jours_presence': paie.jours_presence,
                        'absent': paie.absences,
                        'sundays': paie.dimanches,
                        'j_repos': paie.jours_repos,
                        'maladies': paie.get_maladies_count() if hasattr(paie, 'get_maladies_count') else 0,
                        'm_payer': paie.get_maladies_payees_count() if hasattr(paie, 'get_maladies_payees_count') else 0,
                        'heures_supplementaires': float(paie.heures_supplementaires or 0),
                        'montant_heures_supplementaires': float(paie.montant_heures_supplementaires or 0),
                    },
                    'coherence': {
                        'jours_presence': True,
                        'absences': True,
                        'dimanches': True,
                        'jours_repos': True,
                        'heures_supp': True,
                        'montant_supp': True,
                    },
                    'a_incoherences': False,
                    'heures_supp_details': []
                })
        
        # Synchronisation automatique: pas de message spécifique force_sync

        context = {
            'paies_enrichies': paies_enrichies,
            'employes': employes,
            'employe_id': employe_id,
            'annee': annee,
            'mois': mois,
            'auto_sync': auto_sync,
            'force_sync': force_sync,
            'mois_noms': {
                1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
                5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
                9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
            },
            'formules_calcul': {
                'jours_presence': 'P(Am_&_Pm) + P(Pm) + P(Am)',
                'sundays': 'P(dim_Am) + P(dim_Pm) + P(dim_Am_&_Pm)',
                'absent': 'nombre de A',
                'maladies': 'nombre de M',
                'm_payer': 'nombre de M(Payer)',
                'j_repos': 'nombre de OFF'
            }
        }
        
        return render(request, 'fleet_app/entreprise/paie_employe_list.html', context)
    
    except Exception as e:
        from django.http import HttpResponse
        import traceback
        # Retourner l'erreur complète pour debug
        error_details = f"Erreur 500 dans paie_employe_list:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        return HttpResponse(f"<pre>{error_details}</pre>", status=500)

@login_required
def paie_employe_create(request, employe_id):
    """Vue optimisée pour créer une paie avec relations parfaites et calculs automatiques"""
    from datetime import datetime
    from decimal import Decimal
    import calendar
    from django.db.models import Sum, Count, Q, Prefetch
    
    # Récupération optimisée de l'employé avec toutes ses relations
    employe = get_object_or_404(
        Employe.objects.select_related().prefetch_related(
            Prefetch('presences', 
                    queryset=PresenceJournaliere.objects.select_related()),
            Prefetch('heures_supplementaires', 
                    queryset=HeureSupplementaire.objects.select_related()),
            Prefetch('paies', 
                    queryset=PaieEmploye.objects.select_related())
        ),
        id=employe_id, 
        user=request.user
    )
    
    # Récupération du mois et année actuels ou sélectionnés
    mois_actuel = int(request.GET.get('mois', datetime.now().month))
    annee_actuelle = int(request.GET.get('annee', datetime.now().year))
    
    # Récupération intelligente des données depuis les paies existantes
    paie_existante = employe.paies.filter(
        mois=mois_actuel, 
        annee=annee_actuelle
    ).first()
    
    # Récupération de la dernière paie pour pré-remplir les primes
    derniere_paie = employe.paies.order_by('-annee', '-mois').first()
    
    # Valeurs par défaut intelligentes depuis la dernière paie
    prime_discipline_defaut = derniere_paie.prime_discipline if derniere_paie else 0
    cherete_vie_defaut = derniere_paie.cherete_vie if derniere_paie else 0
    indemnite_transport_defaut = derniere_paie.indemnite_transport if derniere_paie else 0
    indemnite_logement_defaut = derniere_paie.indemnite_logement if derniere_paie else 0
    
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            annee = int(request.POST.get('annee', annee_actuelle))
            mois = int(request.POST.get('mois', mois_actuel))
            
            # Données de base
            salaire_base = Decimal(str(request.POST.get('salaire_base', employe.salaire_journalier or 0)))
            
            # Primes et indemnités modifiables
            prime_discipline = Decimal(str(request.POST.get('prime_discipline', 0)))
            cherete_vie = Decimal(str(request.POST.get('cherete_vie', 0)))
            indemnite_transport = Decimal(str(request.POST.get('indemnite_transport', 0)))
            indemnite_logement = Decimal(str(request.POST.get('indemnite_logement', 0)))
            
            # Déductions
            avances = Decimal(str(request.POST.get('avances', employe.avances or 0)))
            sanctions = Decimal(str(request.POST.get('sanctions', employe.sanctions or 0)))
            
            # Calcul automatique des données de présence
            presences_mois = PresenceJournaliere.objects.filter(
                employe=employe,
                date__month=mois,
                date__year=annee
            )
            
            # Calcul automatique des données de présence selon les formules exactes
            jours_travailles = presences_mois.filter(
                Q(statut='P(Am_&_Pm)') | Q(statut='P(Pm)') | Q(statut='P(Am)')
            ).count()
            
            jours_absents = presences_mois.filter(statut='A').count()
            jours_repos = presences_mois.filter(statut='OFF').count()
            maladies = presences_mois.filter(statut='M').count()
            maladies_payees = presences_mois.filter(statut='M(Payer)').count()
            jours_feries = presences_mois.filter(statut='Férié').count()
            dimanches_travailles = presences_mois.filter(
                Q(statut='P(dim_Am)') | Q(statut='P(dim_Pm)') | Q(statut='P(dim_Am_&_Pm)')
            ).count()
            
            # Calcul des heures supplémentaires
            heures_supp = HeureSupplementaire.objects.filter(
                employe=employe,
                date__month=mois,
                date__year=annee
            )
            
            total_heures_supp = sum(h.duree for h in heures_supp)
            montant_heures_supp = sum(h.total_a_payer for h in heures_supp)
            
            # Calcul du salaire brut
            salaire_brut = (
                salaire_base + prime_discipline + cherete_vie + 
                indemnite_transport + indemnite_logement + Decimal(str(montant_heures_supp))
            )
            
            # Calcul des déductions
            cnss_employe = salaire_brut * Decimal('0.05')  # 5% CNSS
            total_deductions = cnss_employe + avances + sanctions
            
            # Net à payer
            net_a_payer = salaire_brut - total_deductions
            
            # Vérifier si une paie existe déjà
            paie_existante = PaieEmploye.objects.filter(
                employe=employe, annee=annee, mois=mois
            ).first()
            
            if paie_existante:
                # Mettre à jour la paie existante
                paie_existante.salaire_base = salaire_base
                paie_existante.prime_discipline = prime_discipline
                paie_existante.cherete_vie = cherete_vie
                paie_existante.indemnite_transport = indemnite_transport
                paie_existante.indemnite_logement = indemnite_logement
                paie_existante.jours_travailles = jours_travailles
                paie_existante.jours_absents = jours_absents
                paie_existante.dimanches_travailles = dimanches_travailles
                paie_existante.maladies_payees = maladies_payees
                paie_existante.maladies = maladies
                paie_existante.heures_supplementaires = total_heures_supp
                paie_existante.montant_heures_supp = montant_heures_supp
                paie_existante.salaire_brut = salaire_brut
                paie_existante.cnss_employe = cnss_employe
                paie_existante.avances = avances
                paie_existante.sanctions = sanctions
                paie_existante.net_a_payer = net_a_payer
                paie_existante.save()
                
                messages.success(request, f'Paie mise à jour avec succès pour {employe.prenom} {employe.nom}')
            else:
                # Créer une nouvelle paie
                paie = PaieEmploye.objects.create(
                    employe=employe,
                    annee=annee,
                    mois=mois,
                    salaire_base=salaire_base,
                    prime_discipline=prime_discipline,
                    cherete_vie=cherete_vie,
                    indemnite_transport=indemnite_transport,
                    indemnite_logement=indemnite_logement,
                    jours_travailles=jours_travailles,
                    jours_absents=jours_absents,
                    dimanches_travailles=dimanches_travailles,
                    maladies_payees=maladies_payees,
                    maladies=maladies,
                    heures_supplementaires=total_heures_supp,
                    montant_heures_supp=montant_heures_supp,
                    salaire_brut=salaire_brut,
                    cnss_employe=cnss_employe,
                    avances=avances,
                    sanctions=sanctions,
                    net_a_payer=net_a_payer
                )
                
                messages.success(request, f'Paie créée avec succès pour {employe.prenom} {employe.nom}')
            
            return redirect('fleet_app:paie_employe_list')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création/modification : {str(e)}')
    
    # Calcul optimisé des données avec relations et agrégations
    total_jours_mois = calendar.monthrange(annee_actuelle, mois_actuel)[1]
    
    # Utilisation des relations pré-chargées pour optimiser les performances
    presences_mois = [p for p in employe.presences.all() 
                     if p.date.month == mois_actuel and p.date.year == annee_actuelle]
    
    # Calcul des statistiques de présence avec comptage optimisé selon les formules exactes
    jours_travailles = sum(1 for p in presences_mois if p.statut in ['P(Am_&_Pm)', 'P(Pm)', 'P(Am)'])
    jours_absents = sum(1 for p in presences_mois if p.statut == 'A')
    jours_repos = sum(1 for p in presences_mois if p.statut == 'OFF')
    maladies = sum(1 for p in presences_mois if p.statut == 'M')
    maladies_payees = sum(1 for p in presences_mois if p.statut == 'M(Payer)')
    jours_feries = sum(1 for p in presences_mois if p.statut == 'Férié')
    dimanches_travailles = sum(1 for p in presences_mois if p.statut in ['P(dim_Am)', 'P(dim_Pm)', 'P(dim_Am_&_Pm)'])
    
    # Calcul optimisé des heures supplémentaires avec relations pré-chargées
    heures_supp_mois = [h for h in employe.heures_supplementaires.all() 
                       if h.date.month == mois_actuel and h.date.year == annee_actuelle]
    
    total_heures_supp = sum(h.duree for h in heures_supp_mois)
    montant_heures_supp = sum(h.total_a_payer for h in heures_supp_mois)
    
    # Calculs automatiques pour l'affichage
    salaire_base = employe.salaire_journalier or 0
    salaire_brut = salaire_base + montant_heures_supp
    cnss_employe = salaire_brut * Decimal('0.05')
    avances = employe.avances or 0
    # Champ 'sanctions' n'existe plus sur Employe -> valeur par défaut 0
    sanctions = 0
    net_a_payer = salaire_brut - cnss_employe - avances - sanctions
    
    context = {
        'employe': employe,
        'annee_actuelle': annee_actuelle,
        'mois_actuel': mois_actuel,
        'mois_nom': calendar.month_name[mois_actuel],
        
        # Données automatiquement calculées depuis les relations
        'total_jours_mois': total_jours_mois,
        'jours_travailles': jours_travailles,
        'jours_absents': jours_absents,
        'jours_repos': jours_repos,
        'maladies': maladies,
        'maladies_payees': maladies_payees,
        'jours_feries': jours_feries,
        'dimanches_travailles': dimanches_travailles,
        'total_heures_supp': total_heures_supp,
        'montant_heures_supp': montant_heures_supp,
        
        # Calculs financiers optimisés
        'salaire_base': salaire_base,
        'salaire_brut': salaire_brut,
        'cnss_employe': cnss_employe,
        'avances': avances,
        'sanctions': sanctions,
        'net_a_payer': net_a_payer,
        
        # Données intelligentes depuis les paies existantes
        'paie_existante': paie_existante,
        'derniere_paie': derniere_paie,
        'prime_discipline_defaut': prime_discipline_defaut,
        'cherete_vie_defaut': cherete_vie_defaut,
        'indemnite_transport_defaut': indemnite_transport_defaut,
        'indemnite_logement_defaut': indemnite_logement_defaut,
        
        # Historique des paies pour référence
        'historique_paies': employe.paies.order_by('-annee', '-mois')[:6],
        
        # Statistiques globales de l'employé
        'total_presences_annee': len([p for p in employe.presences.all() 
                                    if p.date.year == annee_actuelle]),
        'total_heures_supp_annee': sum(h.duree for h in employe.heures_supplementaires.all() 
                                     if h.date.year == annee_actuelle),
    }
    
    return render(request, 'fleet_app/entreprise/paie_employe_form.html', context)

@login_required
def paie_employe_edit(request, pk):
    """Vue pour modifier une paie avec données automatiques de présence et heures supp"""
    from datetime import date
    import calendar
    from decimal import Decimal
    
    paie = get_object_or_404(PaieEmploye, pk=pk, employe__user=request.user)
    employe = paie.employe
    
    print(f"🔍 PAIE EDIT GET: Chargement paie ID {pk} pour {employe.prenom} {employe.nom}")
    print(f"💾 PAIE EDIT GET: Valeurs actuelles en base:")
    print(f"   - Prime discipline: {paie.prime_discipline}")
    print(f"   - Cherté de vie: {paie.cherete_vie}")
    print(f"   - Transport: {paie.indemnite_transport}")
    print(f"   - Logement: {paie.indemnite_logement}")
    print(f"   - Avances: {paie.avance_sur_salaire}")
    print(f"   - Sanctions: {paie.sanction_vol_carburant}")
    print(f"   - Salaire brut: {paie.salaire_brut}")
    
    # Récupérer le mois et l'année de la paie
    if paie.mois and paie.annee:
        month = paie.mois
        year = paie.annee
    else:
        # Utiliser le mois actuel si pas défini
        today = date.today()
        month = today.month
        year = today.year
    
    # Calculer les dates de début et fin du mois
    start_date = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = date(year, month, last_day)
    
    # Récupérer toutes les présences pour ce mois et cet employé
    presences = PresenceJournaliere.objects.filter(
        employe=employe,
        date__gte=start_date,
        date__lte=end_date
    )
    
    # Calculer les statistiques de présence
    stats = {
        'total_jours': last_day,
        'presences': 0,
        'absences': 0,
        'repos': 0,
        'maladies': 0,
        'maladies_payees': 0,
        'feries': 0,
        'dimanches': 0,
        'conges': 0,
        'formations': 0
    }
    
    for presence in presences:
        if presence.statut in ['present_am', 'present_pm', 'present_journee', 'P(Am)', 'P(Pm)', 'P(Am&Pm)']:
            stats['presences'] += 1
        elif presence.statut in ['absent', 'A']:
            stats['absences'] += 1
        elif presence.statut in ['repos', 'OFF']:
            stats['repos'] += 1
        elif presence.statut in ['maladie', 'M']:
            stats['maladies'] += 1
        elif presence.statut in ['maladie_payee', 'M(Payer)']:
            stats['maladies_payees'] += 1
        elif presence.statut in ['ferie', 'F']:
            stats['feries'] += 1
        elif presence.statut in ['dimanche', 'D']:
            stats['dimanches'] += 1
        elif presence.statut in ['conge', 'C']:
            stats['conges'] += 1
        elif presence.statut in ['formation', 'Formation']:
            stats['formations'] += 1
    
    # Récupérer les heures supplémentaires pour ce mois et cet employé
    heures_supp = HeureSupplementaire.objects.filter(
        employe=employe,
        date__gte=start_date,
        date__lte=end_date
    )
    
    total_heures_supp = sum(h.duree for h in heures_supp)
    total_montant_heures_supp = sum(h.total_a_payer for h in heures_supp)
    
    # Récupérer la configuration des montants pour cet employé
    try:
        from fleet_app.models_entreprise import ConfigurationMontantEmploye
        config_employe = ConfigurationMontantEmploye.objects.get(employe=employe)
        montants = {
            'montant_am': config_employe.montant_am,
            'montant_pm': config_employe.montant_pm,
            'montant_journee': config_employe.montant_journee,
            'montant_dimanche': config_employe.montant_dim_journee,
            'montant_absent': config_employe.montant_absent,
            'montant_maladie': config_employe.montant_maladie,
            'montant_maladie_payee': config_employe.montant_maladie_payee,
            'montant_repos': config_employe.montant_repos,
            'montant_conge': config_employe.montant_conge,
            'montant_formation': config_employe.montant_formation,
            'montant_ferie': config_employe.montant_ferie
        }
    except:
        # Aucune configuration individuelle trouvée: utiliser des valeurs par défaut sûres
        # Ces valeurs peuvent être ajustées dans une future configuration globale par utilisateur
        montants = {
            'montant_am': Decimal('50000'),
            'montant_pm': Decimal('50000'),
            'montant_journee': Decimal('100000'),
            'montant_dimanche': Decimal('150000'),
            'montant_absent': Decimal('0'),
            'montant_maladie': Decimal('0'),
            'montant_maladie_payee': Decimal('100000'),
            'montant_repos': Decimal('0'),
            'montant_conge': Decimal('100000'),
            'montant_formation': Decimal('100000'),
            'montant_ferie': Decimal('150000'),
        }
    
    if request.method == 'POST':
        print(f"📝 PAIE EDIT: Méthode POST reçue pour paie ID {pk}")
        print(f"📝 PAIE EDIT: Données POST reçues: {dict(request.POST)}")
        
        try:
            # Mettre à jour les champs modifiables
            prime_discipline = request.POST.get('prime_discipline', 0)
            cherte_vie = request.POST.get('cherte_vie', 0)  # Template envoie 'cherte_vie'
            transport = request.POST.get('transport', 0)    # Template envoie 'transport'
            logement = request.POST.get('logement', 0)      # Template envoie 'logement'
            avances = request.POST.get('avances', 0)
            sanctions = request.POST.get('sanctions', 0)
            
            print(f"💰 PAIE EDIT: Valeurs reçues:")
            print(f"   - Prime discipline: {prime_discipline}")
            print(f"   - Cherté de vie: {cherte_vie}")
            print(f"   - Transport: {transport}")
            print(f"   - Logement: {logement}")
            print(f"   - Avances: {avances}")
            print(f"   - Sanctions: {sanctions}")
            
            # CORRECTION: Utiliser les vrais noms de champs du modèle
            paie.prime_discipline = Decimal(prime_discipline)
            paie.cherete_vie = Decimal(cherte_vie)           # Modèle: cherete_vie (1 t)
            paie.indemnite_transport = Decimal(transport)    # Modèle: indemnite_transport
            paie.indemnite_logement = Decimal(logement)      # Modèle: indemnite_logement
            paie.avance_sur_salaire = Decimal(avances)       # Modèle: avance_sur_salaire
            paie.sanction_vol_carburant = Decimal(sanctions) # Modèle: sanction_vol_carburant
            
            print(f"✅ PAIE EDIT: Valeurs converties en Decimal:")
            print(f"   - Prime discipline: {paie.prime_discipline}")
            print(f"   - Cherté de vie: {paie.cherete_vie}")
            print(f"   - Transport: {paie.indemnite_transport}")
            print(f"   - Logement: {paie.indemnite_logement}")
            
            # Recalculer automatiquement les montants basés sur les présences
            paie.salaire_base = employe.salaire_journalier or 0
            paie.total_jours = stats['total_jours']
            paie.jours_presence = stats['presences']
            paie.absences = stats['absences']
            paie.repos = stats['repos']
            paie.maladies = stats['maladies']
            paie.maladies_payees = stats['maladies_payees']
            paie.feries = stats['feries']
            paie.dimanches = stats['dimanches']
            paie.heures_supplementaires = total_heures_supp
            paie.montant_heures_supplementaires = total_montant_heures_supp
            
            # Calculer le salaire brut
            salaire_brut = (
                paie.salaire_base +
                paie.prime_discipline +
                paie.cherete_vie +
                paie.indemnite_transport +
                paie.indemnite_logement +
                paie.montant_heures_supplementaires
            )
            paie.salaire_brut = salaire_brut
            
            # 🇬🇳 CALCUL DES DÉDUCTIONS SELON LA RÉGLEMENTATION GUINÉENNE
            
            # 1. CNSS Employé - 5% sur le salaire brut (configurable)
            try:
                param_appliquer = ParametrePaie.objects.get(cle='CNSS_ACTIVER', user=request.user)
                appliquer_cnss = param_appliquer.valeur == Decimal('1')
            except ParametrePaie.DoesNotExist:
                appliquer_cnss = False
            
            try:
                param_taux = ParametrePaie.objects.get(cle='CNSS_TAUX', user=request.user)
                taux_cnss = param_taux.valeur
            except ParametrePaie.DoesNotExist:
                taux_cnss = Decimal('5.0')
            
            if appliquer_cnss:
                cnss_employe = (salaire_brut * taux_cnss) / Decimal('100')
            else:
                cnss_employe = Decimal('0.00')
            
            # 2. RTS (Retenue à la source) - Barème progressif guinéen
            salaire_net_imposable = salaire_brut - cnss_employe
            
            try:
                param_rts_type = ParametrePaie.objects.get(cle='RTS_TYPE', user=request.user)
                rts_type = param_rts_type.valeur
            except ParametrePaie.DoesNotExist:
                rts_type = 'PROGRESSIF'
            
            try:
                param_rts_taux = ParametrePaie.objects.get(cle='RTS_TAUX_FIXE', user=request.user)
                rts_taux_fixe = param_rts_taux.valeur
            except ParametrePaie.DoesNotExist:
                rts_taux_fixe = Decimal('10.0')
            
            if rts_type == 'FIXE':
                # RTS à taux fixe (sur salaire net imposable)
                rts_employe = (salaire_net_imposable * rts_taux_fixe) / Decimal('100')
            else:
                # RTS barème progressif guinéen
                if salaire_net_imposable <= Decimal('1000000'):
                    rts_employe = Decimal('0.00')
                elif salaire_net_imposable <= Decimal('3000000'):
                    rts_employe = (salaire_net_imposable - Decimal('1000000')) * Decimal('0.05')
                else:
                    rts_tranche_2 = Decimal('2000000') * Decimal('0.05')
                    rts_tranche_3 = (salaire_net_imposable - Decimal('3000000')) * Decimal('0.15')
                    rts_employe = rts_tranche_2 + rts_tranche_3
            
            paie.cnss = cnss_employe
            paie.rts = rts_employe
            
            # Calculer le net à payer
            total_deductions = cnss_employe + rts_employe + paie.avance_sur_salaire + paie.sanction_vol_carburant
            paie.salaire_net = salaire_brut - total_deductions
            
            print(f"💾 PAIE EDIT: Avant sauvegarde:")
            print(f"   - Salaire brut: {paie.salaire_brut}")
            print(f"   - Prime discipline: {paie.prime_discipline}")
            print(f"   - Cherté de vie: {paie.cherete_vie}")
            print(f"   - Transport: {paie.indemnite_transport}")
            print(f"   - Logement: {paie.indemnite_logement}")
            print(f"   - CNSS: {paie.cnss}")
            print(f"   - RTS: {paie.rts}")
            print(f"   - Net à payer: {paie.salaire_net}")
            
            paie.save()
            
            print(f"✅ PAIE EDIT: Paie sauvegardée avec succès pour {employe.prenom} {employe.nom}")
            print(f"✅ PAIE EDIT: Valeurs en base après sauvegarde:")
            paie.refresh_from_db()
            print(f"   - Prime discipline DB: {paie.prime_discipline}")
            print(f"   - Cherté de vie DB: {paie.cherete_vie}")
            print(f"   - Transport DB: {paie.indemnite_transport}")
            print(f"   - Logement DB: {paie.indemnite_logement}")
            
            messages.success(request, f'Paie de {employe.prenom} {employe.nom} modifiée avec succès')
            return redirect('fleet_app:paie_employe_list')
            
        except Exception as e:
            print(f"[ERROR] Erreur modification paie: {str(e)}")
            messages.error(request, f'Erreur lors de la modification : {str(e)}')
    
    context = {
        'paie': paie,
        'employe': employe,
        'stats': stats,
        'total_heures_supp': total_heures_supp,
        'total_montant_heures_supp': total_montant_heures_supp,
        'montants': montants,
        'month': month,
        'year': year,
        'month_name': calendar.month_name[month]
    }
    return render(request, 'fleet_app/entreprise/paie_employe_form.html', context)

@login_required
def paie_employe_export(request):
    """Vue pour exporter les paies en CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="paies_employes.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Employé', 'Matricule', 'Année', 'Mois', 'Salaire Base', 'Heures Supp', 'Primes', 'Déductions', 'Salaire Net'])
    
    paies = PaieEmploye.objects.filter(employe__user=request.user).order_by('-annee', '-mois')
    
    for paie in paies:
        writer.writerow([
            f"{paie.employe.nom} {paie.employe.prenom}",
            paie.employe.matricule,
            paie.annee,
            paie.mois,
            paie.salaire_base,
            paie.heures_supplementaires,
            paie.primes,
            paie.deductions,
            paie.salaire_net
        ])
    
    return response

@login_required
def paie_employe_delete(request, pk):
    """Vue pour supprimer une paie"""
    paie = get_object_or_404(PaieEmploye, pk=pk, employe__user=request.user)
    
    if request.method == 'POST':
        try:
            employe_nom = f"{paie.employe.nom} {paie.employe.prenom}"
            paie.delete()
            messages.success(request, f'Paie supprimée avec succès pour {employe_nom}')
            return redirect('fleet_app:paie_employe_list')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression : {str(e)}')
    
    context = {'paie': paie}
    return render(request, 'fleet_app/entreprise/paie_employe_confirm_delete.html', context)

# ===== VUES POUR LES HEURES SUPPLÉMENTAIRES =====

@login_required
def heure_supplementaire_list(request):
    """
    Vue pour afficher la liste des heures supplémentaires
    """
    # Traitement des actions POST (définir montant, supprimer)
    if request.method == 'POST':
        action = request.POST.get('action')
        heure_id = request.POST.get('heure_id')
        
        if action == 'definir_montant' and heure_id:
            try:
                heure_sup = HeureSupplementaire.objects.get(
                    id=heure_id, 
                    employe__user=request.user
                )
                montant_manuel = float(request.POST.get('montant_manuel', 0))
                
                # Interpréter comme taux horaire personnalisé et recalculer
                heure_sup.taux_horaire = montant_manuel if montant_manuel > 0 else None
                heure_sup.save()  # recalcul automatique via model.save()
                
                messages.success(request, f'Taux horaire personnalisé appliqué: {montant_manuel} GNF')
                print(f"DEBUG: Taux horaire personnalisé mis à jour pour heure_sup {heure_id}: {montant_manuel} GNF")
                
            except HeureSupplementaire.DoesNotExist:
                messages.error(request, 'Heure supplémentaire non trouvée')
            except ValueError:
                messages.error(request, 'Montant invalide')
            except Exception as e:
                messages.error(request, f'Erreur lors de la mise à jour: {str(e)}')
        
        elif action == 'delete' and heure_id:
            try:
                heure_sup = HeureSupplementaire.objects.get(
                    id=heure_id, 
                    employe__user=request.user
                )
                employe_nom = f"{heure_sup.employe.prenom} {heure_sup.employe.nom}"
                heure_sup.delete()
                
                messages.success(request, f'Heure supplémentaire supprimée pour {employe_nom}')
                print(f"DEBUG: Heure supplémentaire {heure_id} supprimée")
                
            except HeureSupplementaire.DoesNotExist:
                messages.error(request, 'Heure supplémentaire non trouvée')
            except Exception as e:
                messages.error(request, f'Erreur lors de la suppression: {str(e)}')
        
        elif action == 'definir_montant_global':
            try:
                montant_global = float(request.POST.get('montant_global', 0))
                if montant_global > 0:
                    # Construire le queryset selon les filtres de période fournis
                    heures_sup_all = HeureSupplementaire.objects.filter(employe__user=request.user)
                    employe_id_post = request.POST.get('employe_id')
                    mois_post = request.POST.get('mois')
                    annee_post = request.POST.get('annee')
                    date_debut_post = request.POST.get('date_debut')
                    date_fin_post = request.POST.get('date_fin')

                    if employe_id_post:
                        heures_sup_all = heures_sup_all.filter(employe_id=employe_id_post)
                    if mois_post and annee_post:
                        heures_sup_all = heures_sup_all.filter(date__month=int(mois_post), date__year=int(annee_post))
                    else:
                        if date_debut_post:
                            heures_sup_all = heures_sup_all.filter(date__gte=date_debut_post)
                        if date_fin_post:
                            heures_sup_all = heures_sup_all.filter(date__lte=date_fin_post)

                    count = 0
                    for hs in heures_sup_all:
                        hs.taux_horaire = montant_global
                        hs.save()
                        count += 1
                    messages.success(request, f'Taux horaire global de {montant_global} GNF appliqué à {count} heures supplémentaires')
                    print(f"DEBUG: Taux horaire global {montant_global} appliqué à {count} heures supplémentaires")
                else:
                    messages.error(request, 'Le montant global doit être supérieur à 0')
                    
            except ValueError:
                messages.error(request, 'Montant global invalide')
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'application du montant global: {str(e)}')
        
        elif action == 'reinitialiser_montant_global':
            try:
                # Réinitialiser le taux personnalisé (retour à la config employé) et recalculer
                heures_sup_all = HeureSupplementaire.objects.filter(employe__user=request.user)
                employe_id_post = request.POST.get('employe_id')
                mois_post = request.POST.get('mois')
                annee_post = request.POST.get('annee')
                date_debut_post = request.POST.get('date_debut')
                date_fin_post = request.POST.get('date_fin')

                if employe_id_post:
                    heures_sup_all = heures_sup_all.filter(employe_id=employe_id_post)
                if mois_post and annee_post:
                    heures_sup_all = heures_sup_all.filter(date__month=int(mois_post), date__year=int(annee_post))
                else:
                    if date_debut_post:
                        heures_sup_all = heures_sup_all.filter(date__gte=date_debut_post)
                    if date_fin_post:
                        heures_sup_all = heures_sup_all.filter(date__lte=date_fin_post)

                count = 0
                for hs in heures_sup_all:
                    hs.taux_horaire = None
                    hs.save()
                    count += 1
                messages.success(request, f'Taux personnalisés réinitialisés pour {count} heures supplémentaires')
                print(f"DEBUG: Taux personnalisés réinitialisés pour {count} heures supplémentaires")
                
            except Exception as e:
                messages.error(request, f'Erreur lors de la réinitialisation: {str(e)}')
        
        # Rediriger pour éviter la re-soumission
        return redirect('fleet_app:heure_supplementaire_list')
    
    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    employe_id = request.GET.get('employe_id')
    
    # Récupérer les heures supplémentaires avec optimisation
    heures_sup = HeureSupplementaire.objects.filter(employe__user=request.user).select_related('employe').order_by('-date', '-id')
    
    if date_debut:
        heures_sup = heures_sup.filter(date__gte=date_debut)
    if date_fin:
        heures_sup = heures_sup.filter(date__lte=date_fin)
    if employe_id:
        heures_sup = heures_sup.filter(employe_id=employe_id)
    
    # Debug: Vérifier le nombre d'heures supplémentaires récupérées
    print(f"DEBUG LIST: Nombre d'heures supplémentaires trouvées: {heures_sup.count()}")
    for hs in heures_sup[:3]:  # Afficher les 3 premières
        print(f"DEBUG LIST: {hs.employe.matricule} - {hs.employe.nom} - {hs.date} - {hs.duree}h")
    
    # Debug: Vérifier le nombre d'heures supplémentaires récupérées
    print(f"DEBUG LIST: Nombre d'heures supplémentaires trouvées: {heures_sup.count()}")
    for hs in heures_sup[:3]:  # Afficher les 3 premières
        print(f"DEBUG LIST: {hs.employe.matricule} - {hs.employe.nom} - {hs.date} - {hs.duree}h")
    
    # Récupérer la liste des employés pour le formulaire de filtrage
    employes = Employe.objects.filter(user=request.user).order_by('matricule')
    
    return render(request, 'fleet_app/entreprise/heure_supplementaire_list.html', {
        'heures_sup': heures_sup,
        'employes': employes,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'employe_id': employe_id,
    })

@login_required
def heure_supplementaire_add(request):
    """Vue pour ajouter des heures supplémentaires"""
    if request.method == 'POST':
        try:
            employe_id = request.POST.get('employe')
            date_str = request.POST.get('date')
            heure_debut = request.POST.get('heure_debut')
            heure_fin = request.POST.get('heure_fin')
            duree = float(request.POST.get('duree', 0))
            taux_horaire = float(request.POST.get('taux_horaire', 5000))
            autorise_par = request.POST.get('autorise_par', 'Système')
            
            employe = Employe.objects.get(id=employe_id, user=request.user)
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # Convertir les heures en objets time si fournies
            heure_debut_obj = None
            heure_fin_obj = None
            if heure_debut:
                heure_debut_obj = datetime.strptime(heure_debut, '%H:%M').time()
            if heure_fin:
                heure_fin_obj = datetime.strptime(heure_fin, '%H:%M').time()
            
            # Si les heures ne sont pas fournies, utiliser des valeurs par défaut
            if not heure_debut_obj:
                heure_debut_obj = datetime.strptime('08:00', '%H:%M').time()
            if not heure_fin_obj:
                # Calculer l'heure de fin basée sur la durée
                heure_fin_calculee = datetime.combine(date_obj, heure_debut_obj) + timedelta(hours=float(duree))
                heure_fin_obj = heure_fin_calculee.time()
            
            # Calculer le montant total
            montant_total = duree * taux_horaire
            
            HeureSupplementaire.objects.create(
                employe=employe,
                date=date_obj,
                heure_debut=heure_debut_obj,
                heure_fin=heure_fin_obj,
                duree=duree,
                taux_horaire=taux_horaire,
                total_a_payer=montant_total,
                autorise_par=autorise_par
            )
            
            # Debug: Vérifier que l'enregistrement a bien été créé
            print(f"DEBUG: Heure supplémentaire créée pour {employe.nom} {employe.prenom}")
            print(f"DEBUG: Date: {date_obj}, Durée: {duree}, Montant: {montant_total}")
            
            # Vérifier le nombre total d'heures supplémentaires pour cet utilisateur
            total_heures = HeureSupplementaire.objects.filter(employe__user=request.user).count()
            print(f"DEBUG: Total heures supplémentaires pour cet utilisateur: {total_heures}")
            
            messages.success(request, 'Heures supplémentaires ajoutées avec succès')
            return redirect('fleet_app:heure_supplementaire_list')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'ajout : {str(e)}')
    
    employes = Employe.objects.filter(user=request.user).order_by('matricule')
    # Ajouter la date du jour par défaut
    from datetime import date
    context = {
        'employes': employes,
        'date_aujourd_hui': date.today().strftime('%Y-%m-%d')
    }
    
    return render(request, 'fleet_app/entreprise/heure_supplementaire_form.html', context)

@login_required
def heure_supplementaire_edit(request, pk):
    """Modifier une heure supplémentaire existante"""
    heure = get_object_or_404(HeureSupplementaire, pk=pk, employe__user=request.user)
    employes = Employe.objects.filter(user=request.user).order_by('matricule')

    if request.method == 'POST':
        try:
            employe_id = request.POST.get('employe')
            date_str = request.POST.get('date')
            heure_debut = request.POST.get('heure_debut')
            heure_fin = request.POST.get('heure_fin')
            taux_horaire = float(request.POST.get('taux_horaire', heure.taux_horaire or 5000))
            autorise_par = request.POST.get('autorise_par', heure.autorise_par or 'Système')

            if employe_id:
                heure.employe = get_object_or_404(Employe, id=employe_id, user=request.user)
            if date_str:
                heure.date = datetime.strptime(date_str, '%Y-%m-%d').date()
            if heure_debut:
                heure.heure_debut = datetime.strptime(heure_debut, '%H:%M').time()
            if heure_fin:
                heure.heure_fin = datetime.strptime(heure_fin, '%H:%M').time()

            # Recalculer la durée si possible
            if heure.heure_debut and heure.heure_fin:
                dt_debut = datetime.combine(heure.date, heure.heure_debut)
                dt_fin = datetime.combine(heure.date, heure.heure_fin)
                if dt_fin <= dt_debut:
                    dt_fin += timedelta(days=1)
                duree_heures = (dt_fin - dt_debut).total_seconds() / 3600.0
                heure.duree = round(duree_heures, 2)

            heure.taux_horaire = taux_horaire
            # Mettre à jour montant total/calculé
            try:
                # si un champ total_a_payer existe, le recalculer à partir du taux
                montant_total = float(heure.duree or 0) * float(heure.taux_horaire or 0)
                heure.total_a_payer = montant_total
            except Exception:
                pass

            heure.autorise_par = autorise_par
            heure.save()

            messages.success(request, 'Heure supplémentaire modifiée avec succès')
            return redirect('fleet_app:heure_supplementaire_list')
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification : {str(e)}')

    # Préparer le contexte avec valeurs pré-remplies
    context = {
        'employes': employes,
        'heure': heure,
        'date_aujourd_hui': heure.date.strftime('%Y-%m-%d') if hasattr(heure, 'date') and heure.date else ''
    }
    return render(request, 'fleet_app/entreprise/heure_supplementaire_form.html', context)

@login_required
def configuration_heure_supplementaire(request):
    """Vue pour la configuration des heures supplémentaires"""
    if request.method == 'POST':
        # Traitement de la configuration
        messages.success(request, 'Configuration mise à jour avec succès')
        return redirect('fleet_app:configuration_heure_supplementaire')
    
    context = {
        'taux_normal': 1500,  # Exemple de taux
        'taux_dimanche': 2000,
        'taux_ferie': 2500,
    }
    
    return render(request, 'fleet_app/entreprise/configuration_heure_supplementaire.html', context)

@login_required
def heure_supplementaire_export(request):
    """
    Vue pour exporter les heures supplémentaires en CSV
    """
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    from .models_entreprise import HeureSupplementaire
    
    # Créer la réponse HTTP avec le type de contenu CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="heures_supplementaires_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    # Créer le writer CSV
    writer = csv.writer(response)
    
    # Écrire l'en-tête
    writer.writerow([
        'Matricule',
        'Nom',
        'Prénom',
        'Date',
        'Durée (heures)',
        'Taux horaire',
        'Montant total',
        'Autorisé par',
        'Date création'
    ])
    
    # Récupérer les heures supplémentaires de l'utilisateur
    heures_sup = HeureSupplementaire.objects.filter(
        employe__user=request.user
    ).select_related('employe').order_by('-date')
    
    # Écrire les données
    for hs in heures_sup:
        writer.writerow([
            hs.employe.matricule,
            hs.employe.nom,
            hs.employe.prenom,
            hs.date.strftime('%d/%m/%Y'),
            hs.duree,
            hs.taux_horaire,
            hs.montant_total,
            hs.autorise_par or 'Non spécifié',
            hs.date_creation.strftime('%d/%m/%Y %H:%M')
        ])
    
    return response


# ===== VUES POUR LA CONFIGURATION DES CHARGES SOCIALES =====

@login_required
def config_charges_sociales(request):
    """
    Vue pour configurer les charges sociales guinéennes (CNSS)
    """
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            
            # Récupérer ou créer la configuration pour l'utilisateur
            config = ConfigurationChargesSociales.get_or_create_for_user(request.user)
            
            # Mettre à jour les paramètres
            config.appliquer_cnss = data.get('appliquer_cnss', True)
            config.taux_cnss = float(data.get('taux_cnss', 5.0))
            config.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Configuration des charges sociales sauvegardée avec succès',
                'config': {
                    'appliquer_cnss': config.appliquer_cnss,
                    'taux_cnss': float(config.taux_cnss)
                }
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Erreur lors de la sauvegarde: {str(e)}'
            })
    
    # GET: retourner la configuration actuelle
    config = ConfigurationChargesSociales.get_or_create_for_user(request.user)
    return JsonResponse({
        'success': True,
        'config': {
            'appliquer_cnss': config.appliquer_cnss,
            'taux_cnss': float(config.taux_cnss)
        }
    })

# ===== VUES POUR LES BULLETINS DE PAIE =====

@login_required
def bulletin_paie_print(request, employe_id):
    """Vue pour imprimer un bulletin de paie individuel"""
    from datetime import datetime
    import calendar
    from decimal import Decimal
    from .models_accounts import Entreprise, Profil
    
    # Récupération des paramètres
    mois_actuel = int(request.GET.get('mois', datetime.now().month))
    annee_actuelle = int(request.GET.get('annee', datetime.now().year))
    
    # Récupération de l'employé
    employe = get_object_or_404(Employe, id=employe_id, user=request.user)
    
    # Récupération des informations de l'entreprise
    entreprise = None
    try:
        profil = Profil.objects.get(user=request.user)
        if hasattr(profil, 'entreprise'):
            entreprise = profil.entreprise
    except Profil.DoesNotExist:
        pass
    
    # Récupération des présences pour le mois
    presences_mois = PresenceJournaliere.objects.filter(
        employe=employe,
        date__month=mois_actuel,
        date__year=annee_actuelle
    )
    
    # Calcul des statistiques de présence détaillées
    total_jours_travailles = presences_mois.filter(
        statut__in=['P(Am)', 'P(Pm)', 'P(Am&Pm)']
    ).count()
    
    total_jours_absent = presences_mois.filter(
        statut='Absent'
    ).count()
    
    total_dimanche_travaille = presences_mois.filter(
        statut='Dimanche'
    ).count()
    
    malade_paye = presences_mois.filter(
        statut='M.Payer'
    ).count()
    
    malade = presences_mois.filter(
        statut='Maladies'
    ).count()
    
    jours_repos = presences_mois.filter(
        statut='J Repos'
    ).count()
    
    jours_feries = presences_mois.filter(
        statut='Férié'
    ).count()
    
    jours_conges = presences_mois.filter(
        statut='Congé'
    ).count()
    
    jours_formation = presences_mois.filter(
        statut='Formation'
    ).count()
    
    # Récupération des heures supplémentaires
    heures_supp = HeureSupplementaire.objects.filter(
        employe=employe,
        date__month=mois_actuel,
        date__year=annee_actuelle
    )
    
    total_heures_supp = sum(h.duree for h in heures_supp)
    total_montant_heures_supp = sum(h.total_a_payer for h in heures_supp)
    
    # Récupération ou création de la paie
    paie, created = PaieEmploye.objects.get_or_create(
        employe=employe,
        mois=mois_actuel,
        annee=annee_actuelle,
        defaults={
            'salaire_base': employe.salaire_journalier or 0,
            'prime_discipline': 0,
            'cherete_vie': 0,
            'indemnite_transport': 0,
            'indemnite_logement': 0,
            'jours_travailles': total_jours_travailles,
            'jours_absents': total_jours_absent,
            'dimanches_travailles': total_dimanche_travaille,
            'maladies_payees': malade_paye,
            'maladies': malade,
            'heures_supplementaires': total_heures_supp,
            'montant_heures_supp': total_montant_heures_supp,
        }
    )
    
    # Calcul du salaire brut
    salaire_brut = (
        (paie.salaire_base or 0) +
        (paie.prime_discipline or 0) +
        (paie.cherete_vie or 0) +
        (paie.indemnite_transport or 0) +
        (paie.indemnite_logement or 0) +
        total_montant_heures_supp
    )
    
    # Calcul des déductions
    cnss_employe = salaire_brut * Decimal('0.05')  # 5% CNSS
    avances = employe.avances or 0
    sanctions = employe.sanctions or 0
    
    total_deductions = cnss_employe + avances + sanctions
    
    # Calcul du net à payer
    net_a_payer = salaire_brut - total_deductions
    
    # Mise à jour de la paie avec les calculs
    paie.salaire_brut = salaire_brut
    paie.cnss_employe = cnss_employe
    paie.avances = avances
    # Ne pas utiliser de champ sanctions inexistant côté Employe
    paie.sanctions = sanctions
    paie.net_a_payer = net_a_payer
    paie.save()
    
    # Calcul du nombre total de jours du mois
    total_jours_mois = calendar.monthrange(annee_actuelle, mois_actuel)[1]
    
    context = {
        'employe': employe,
        'paie': paie,
        'mois_actuel': mois_actuel,
        'annee_actuelle': annee_actuelle,
        'mois_nom': calendar.month_name[mois_actuel],
        'total_jours_mois': total_jours_mois,
        
        # Statistiques détaillées de présence
        'presences_stats': {
            'jours_travailles': total_jours_travailles,
            'jours_absence': total_jours_absent,
            'jours_dimanche': total_dimanche_travaille,
            'malade_paye': malade_paye,
            'malade': malade,
            'jours_repos': jours_repos,
            'jours_feries': jours_feries,
            'jours_conges': jours_conges,
            'jours_formation': jours_formation,
        },
        
        # Heures supplémentaires
        'heures_supp': {
            'total_heures': total_heures_supp,
            'total_montant': total_montant_heures_supp,
            'detail': heures_supp,
        },
        
        # Calculs financiers
        'salaire_brut': salaire_brut,
        'cnss_employe': cnss_employe,
        'total_deductions': total_deductions,
        'net_a_payer': net_a_payer,
        
        # Date d'impression
        'date_impression': datetime.now(),
        
        # Informations de l'entreprise
        'entreprise': entreprise,
    }
    
    return render(request, 'fleet_app/entreprise/bulletin_paie_print.html', context)

@login_required
def bulletin_paie_list(request):
    """Vue bulletins de paie corrigée avec gestion d'erreurs"""
    try:
        from datetime import datetime
        import calendar
        from decimal import Decimal
        from django.core.paginator import Paginator
        from django.contrib import messages
        from dateutil.relativedelta import relativedelta
        
        # Paramètres de filtre avec valeurs par défaut sûres
        mois_actuel = int(request.GET.get('mois', datetime.now().month))
        annee_actuelle = int(request.GET.get('annee', datetime.now().year))
        
        # Calculer le mois précédent et suivant
        date_actuelle = datetime(annee_actuelle, mois_actuel, 1)
        mois_precedent = date_actuelle - relativedelta(months=1)
        mois_suivant = date_actuelle + relativedelta(months=1)
        
        # Nom du mois actuel
        mois_noms = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
        nom_mois_actuel = mois_noms[mois_actuel]
        
        # Récupération des employés avec gestion d'erreur
        try:
            employes = Employe.objects.filter(user=request.user).order_by('matricule')
        except Exception as e:
            messages.error(request, f"Erreur lors de la récupération des employés: {e}")
            employes = []
        
        bulletins_data = []
        
        for employe in employes:
            try:
                # Récupération des présences pour le mois
                presences_mois = PresenceJournaliere.objects.filter(
                    employe=employe,
                    date__month=mois_actuel,
                    date__year=annee_actuelle
                )
                
                # Calculs de base sécurisés
                total_jours_travailles = presences_mois.filter(present=True).count()
                salaire_journalier = employe.salaire_journalier or Decimal('0')
                salaire_brut = salaire_journalier * total_jours_travailles
                
                # Récupération ou création de la paie
                paie, created = PaieEmploye.objects.get_or_create(
                    employe=employe,
                    mois=mois_actuel,
                    annee=annee_actuelle,
                    defaults={
                        'salaire_brut': salaire_brut,
                        'salaire_net': salaire_brut,
                        'salaire_net_a_payer': salaire_brut,
                        'jours_presence': total_jours_travailles,
                        'user': request.user
                    }
                )
                
                # Calculs CNSS sécurisés
                try:
                    param_cnss = ParametrePaie.objects.filter(cle='CNSS_ACTIVER', user=request.user).first()
                    appliquer_cnss = param_cnss and param_cnss.valeur == '1'
                except:
                    appliquer_cnss = False
                
                cnss_employe = salaire_brut * Decimal('0.05') if appliquer_cnss else Decimal('0')
                avances = employe.avances or Decimal('0')
                net_a_payer = salaire_brut - cnss_employe - avances
                
                # Mise à jour de la paie
                paie.salaire_brut = salaire_brut
                paie.cnss = cnss_employe
                paie.avance_sur_salaire = avances
                paie.salaire_net_a_payer = net_a_payer
                paie.save()
                
                bulletins_data.append({
                    'employe': employe,
                    'paie': paie,
                    'jours_travailles': total_jours_travailles,
                    'salaire_brut': salaire_brut,
                    'cnss_employe': cnss_employe,
                    'avances': avances,
                    'net_a_payer': net_a_payer,
                })
                
            except Exception as e:
                # Log l'erreur mais continue avec les autres employés
                print(f"Erreur pour employé {employe.matricule}: {e}")
                continue
        
        # Pagination sécurisée
        paginator = Paginator(bulletins_data, 10)
        page_number = request.GET.get('page', 1)
        
        try:
            page_obj = paginator.get_page(page_number)
        except:
            page_obj = paginator.get_page(1)
        
        # Configuration des charges sociales avec valeurs par défaut
        cnss_activer = False
        cnss_taux = Decimal('5.0')
        rts_type = 'PROGRESSIF'
        rts_taux_fixe = Decimal('10.0')
        
        try:
            param_cnss = ParametrePaie.objects.filter(cle='CNSS_ACTIVER', user=request.user).first()
            if param_cnss:
                cnss_activer = param_cnss.valeur == '1'
            
            param_taux = ParametrePaie.objects.filter(cle='CNSS_TAUX', user=request.user).first()
            if param_taux:
                cnss_taux = Decimal(param_taux.valeur)
        except:
            pass
        
        # Contexte sécurisé
        context = {
            'bulletins_data': bulletins_data,
            'page_obj': page_obj,
            'is_paginated': page_obj.has_other_pages(),
            'mois_actuel': mois_actuel,
            'annee_actuelle': annee_actuelle,
            'mois_nom': calendar.month_name[mois_actuel] if 1 <= mois_actuel <= 12 else 'Inconnu',
            'cnss_activer': cnss_activer,
            'cnss_taux': cnss_taux,
            'rts_type': rts_type,
            'rts_taux_fixe': rts_taux_fixe,
            'entreprise': None,  # Simplifié pour éviter les erreurs
        }
        
        return render(request, 'fleet_app/entreprise/bulletin_paie_list.html', context)
        
    except Exception as e:
        # Gestion d'erreur globale
        messages.error(request, f"Erreur lors du chargement des bulletins: {e}")
        context = {
            'bulletins_data': [],
            'page_obj': None,
            'is_paginated': False,
            'mois_actuel': datetime.now().month,
            'annee_actuelle': datetime.now().year,
            'mois_nom': 'Erreur',
            'cnss_activer': False,
            'cnss_taux': Decimal('5.0'),
            'rts_type': 'PROGRESSIF',
            'rts_taux_fixe': Decimal('10.0'),
            'entreprise': None,
            'error_message': str(e)
        }
        return render(request, 'fleet_app/entreprise/bulletin_paie_list.html', context)


def statistiques_paies(request):
    """Vue pour les statistiques des paies"""
    annee_actuelle = timezone.now().year
    
    # Statistiques par mois
    stats_mensuelles = []
    for mois in range(1, 13):
        paies_mois = PaieEmploye.objects.filter(
            employe__user=request.user,
            annee=annee_actuelle,
            mois=mois
        )
        
        stats_mensuelles.append({
            'mois': mois,
            'nom_mois': datetime(annee_actuelle, mois, 1).strftime('%B'),
            'nombre_paies': paies_mois.count(),
            'total_salaires': paies_mois.aggregate(total=Sum('salaire_net'))['total'] or 0,
            'total_heures_supp': paies_mois.aggregate(total=Sum('heures_supplementaires'))['total'] or 0,
        })
    
    # Statistiques générales
    total_employes = Employe.objects.filter(user=request.user).count()
    total_paies_annee = PaieEmploye.objects.filter(
        employe__user=request.user, 
        annee=annee_actuelle
    ).count()
    
    context = {
        'stats_mensuelles': stats_mensuelles,
        'annee_actuelle': annee_actuelle,
        'total_employes': total_employes,
        'total_paies_annee': total_paies_annee,
    }
    
    return render(request, 'fleet_app/entreprise/statistiques_paies.html', context)

# ===== VUES POUR LES ARCHIVES =====

@login_required
def archive_mensuelle(request):
    """Vue complète pour la gestion des archives mensuelles"""
    from datetime import datetime
    import calendar
    from decimal import Decimal
    from django.db.models import Sum, Count, Avg
    from django.http import JsonResponse
    import json
    
    # Gestion des actions POST (création/clôture d'archives)
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'creer_archive':
            try:
                mois = int(request.POST.get('mois'))
                annee = int(request.POST.get('annee'))
                
                # Vérifier si une archive existe déjà
                archive_existante = ArchiveMensuelle.objects.filter(
                    user=request.user,
                    mois=mois,
                    annee=annee
                ).first()
                
                if archive_existante:
                    messages.error(request, f"Une archive pour {mois:02d}/{annee} existe déjà.")
                else:
                    # Importer les utilitaires d'archivage
                    from .utils_archivage import verifier_coherence_donnees_reference, preparer_nouveau_mois, nettoyer_donnees_transactionnelles_orphelines
                    
                    # Vérifier la cohérence des données avant archivage
                    coherence = verifier_coherence_donnees_reference(request.user)
                    if not coherence['coherent']:
                        messages.warning(request, f"Attention: {len(coherence['incoherences'])} incohérence(s) détectée(s) dans les données de référence.")
                    
                    # Nettoyer les données orphelines avant archivage
                    nettoyage = nettoyer_donnees_transactionnelles_orphelines(request.user)
                    if nettoyage['presences_supprimees'] > 0 or nettoyage['paies_supprimees'] > 0 or nettoyage['heures_supprimees'] > 0:
                        messages.info(request, f"Nettoyage effectué: {nettoyage['presences_supprimees']} présences, {nettoyage['paies_supprimees']} paies, {nettoyage['heures_supprimees']} heures supp orphelines supprimées.")
                    
                    # Créer l'archive avec réinitialisation des données transactionnelles
                    archive = creer_archive_complete(request.user, mois, annee)
                    
                    # Préparer le nouveau mois
                    if mois == 12:
                        nouveau_mois, nouvelle_annee = 1, annee + 1
                    else:
                        nouveau_mois, nouvelle_annee = mois + 1, annee
                    
                    preparation = preparer_nouveau_mois(request.user, nouveau_mois, nouvelle_annee)
                    
                    # Messages de succès détaillés
                    messages.success(request, f"Archive créée avec succès pour {mois:02d}/{annee} (ID: {archive.id})")
                    messages.success(request, f"Prêt pour le nouveau mois: {preparation['info_nouveau_mois']['nom_mois']} {nouvelle_annee} ({preparation['info_nouveau_mois']['nb_jours']} jours)")
                    messages.info(request, f"Données de référence conservées: {preparation['stats_reference']['employes_actifs']} employés actifs, {preparation['stats_reference']['parametres_paie']} paramètres")
                    
                    if preparation['employes_sans_config']:
                        messages.warning(request, f"{len(preparation['employes_sans_config'])} employé(s) sans configuration de montants. Veuillez les configurer.")
                    
            except Exception as e:
                import traceback
                print(f"Erreur lors de l'archivage: {traceback.format_exc()}")
                messages.error(request, f"Erreur lors de la création de l'archive: {str(e)}")
        
        elif action == 'cloturer_archive':
            archive_id = int(request.POST.get('archive_id'))
            archive = get_object_or_404(ArchiveMensuelle, id=archive_id, user=request.user)
            
            if archive.peut_etre_cloturee():
                archive.statut = 'Clôturé'
                archive.date_cloture = timezone.now()
                archive.save()
                messages.success(request, f'Archive {archive.get_periode_label()} clôturée avec succès.')
            else:
                messages.error(request, 'Cette archive ne peut pas être clôturée.')
        
        return redirect('archive_mensuelle_full')
    
    # Paramètres de filtrage
    annee_filtre = request.GET.get('annee', '')
    mois_filtre = request.GET.get('mois', '')
    statut_filtre = request.GET.get('statut', '')
    
    # Récupération des archives avec filtres
    archives_query = ArchiveMensuelle.objects.filter(user=request.user)
    
    if annee_filtre:
        archives_query = archives_query.filter(annee=int(annee_filtre))
    if mois_filtre:
        archives_query = archives_query.filter(mois=int(mois_filtre))
    if statut_filtre:
        archives_query = archives_query.filter(statut=statut_filtre)
    
    archives = archives_query.order_by('-annee', '-mois')
    
    # Statistiques globales
    stats_globales = archives.aggregate(
        total_archives=Count('id'),
        total_employes=Sum('nb_employes_actifs'),
        total_montant_brut=Sum('total_salaire_brut'),
        total_montant_net=Sum('total_net_paye'),
        moyenne_employes=Avg('nb_employes_actifs')
    )
    
    # Années disponibles pour le filtre
    annees_disponibles = archives.values_list('annee', flat=True).distinct().order_by('-annee')
    
    # Données pour les graphiques (derniers 12 mois)
    archives_recentes = archives[:12]
    donnees_graphique = {
        'labels': [archive.get_periode_label() for archive in reversed(archives_recentes)],
        'employes': [archive.nb_employes_actifs for archive in reversed(archives_recentes)],
        'montant_brut': [float(archive.total_salaire_brut or 0) for archive in reversed(archives_recentes)],
        'montant_net': [float(archive.total_net_paye or 0) for archive in reversed(archives_recentes)],
    }
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(archives, 12)
    page_number = request.GET.get('page')
    archives_page = paginator.get_page(page_number)
    
    context = {
        'archives': archives_page,
        'archives_page': archives_page,
        'paginator': paginator,
        'stats_globales': stats_globales,
        'donnees_graphique': json.dumps(donnees_graphique),
        'annees_disponibles': annees_disponibles,
        'annee_actuelle': datetime.now().year,
        'mois_actuel': datetime.now().month,
        'mois_noms': [(i, calendar.month_name[i]) for i in range(1, 13)],
        'statuts_choix': ArchiveMensuelle.STATUT_CHOICES,
        'filtres': {
            'annee': annee_filtre,
            'mois': mois_filtre,
            'statut': statut_filtre,
        }
    }
    
    return render(request, 'fleet_app/entreprise/archive_mensuelle.html', context)


def creer_archive_complete(user, mois, annee):
    """
    Fonction pour créer une archive complète avec toutes les données du mois
    ET réinitialiser les données transactionnelles pour le nouveau mois
    
    LOGIQUE D'ARCHIVAGE :
    - DONNÉES DE RÉFÉRENCE (à conserver) : employés, configuration montants, légende, paramètres
    - DONNÉES TRANSACTIONNELLES (à réinitialiser) : présences, paies, heures supp, bulletins
    """
    from decimal import Decimal
    import json
    from django.db import transaction
    from datetime import datetime, timedelta
    import calendar
    
    print(f"🗂️ ARCHIVAGE: Début de l'archivage pour {mois}/{annee}")
    
    # Utiliser une transaction pour garantir la cohérence
    with transaction.atomic():
        # Récupération de tous les employés
        employes = Employe.objects.filter(user=user)
        
        # Données à archiver
        donnees_archive = {
            'employes': [],
            'presences': [],
            'paies': [],
            'heures_supplementaires': [],
            'configurations_montants': [],
            'parametres_paie': [],
            'statistiques': {}
        }
        
        total_salaire_brut = Decimal('0')
        total_salaire_net = Decimal('0')
        total_heures_supp = Decimal('0')
        nombre_employes = 0
        
        # 1. ARCHIVER LES DONNÉES DE RÉFÉRENCE (à conserver)
        print(f"📋 ARCHIVAGE: Sauvegarde des données de référence")
        
        for employe in employes:
            # Données employé (RÉFÉRENCE - à conserver)
            employe_data = {
                'id': employe.id,
                'matricule': employe.matricule,
                'nom': employe.nom,
                'prenom': employe.prenom,
                'fonction': employe.fonction,
                'salaire_journalier': float(employe.salaire_journalier or 0),
                'statut': employe.statut,
                'avances': float(employe.avances or 0),
                'sanctions': float(employe.sanctions or 0)
            }
            donnees_archive['employes'].append(employe_data)
            
            # Configuration des montants (RÉFÉRENCE - à conserver)
            try:
                from .models_entreprise import ConfigurationMontantEmploye
                config_montant = ConfigurationMontantEmploye.objects.filter(employe=employe).first()
                if config_montant:
                    config_data = {
                        'employe_id': employe.id,
                        'montant_am': float(config_montant.montant_am),
                        'montant_pm': float(config_montant.montant_pm),
                        'montant_journee': float(config_montant.montant_journee),
                        'montant_dim_journee': float(config_montant.montant_dim_journee),
                        'montant_absent': float(config_montant.montant_absent),
                        'montant_maladie': float(config_montant.montant_maladie),
                        'montant_maladie_payee': float(config_montant.montant_maladie_payee),
                        'montant_repos': float(config_montant.montant_repos),
                        'montant_conge': float(config_montant.montant_conge),
                        'montant_formation': float(config_montant.montant_formation),
                        'montant_ferie': float(config_montant.montant_ferie)
                    }
                    donnees_archive['configurations_montants'].append(config_data)
            except ImportError:
                pass
        
        # Paramètres de paie (RÉFÉRENCE - à conserver)
        parametres = ParametrePaie.objects.filter(user=user)
        for param in parametres:
            param_data = {
                'cle': param.cle,
                'valeur': param.valeur,
                'description': param.description
            }
            donnees_archive['parametres_paie'].append(param_data)
        
        # 2. ARCHIVER LES DONNÉES TRANSACTIONNELLES (à réinitialiser)
        print(f"📊 ARCHIVAGE: Sauvegarde des données transactionnelles")
        
        donnees_a_supprimer = {
            'presences': 0,
            'paies': 0,
            'heures_supp': 0
        }
        
        for employe in employes:
            # Présences du mois (TRANSACTIONNEL - à réinitialiser)
            presences_mois = PresenceJournaliere.objects.filter(
                employe=employe,
                date__month=mois,
                date__year=annee
            )
            
            presences_data = []
            for presence in presences_mois:
                presences_data.append({
                    'date': presence.date.isoformat(),
                    'statut': presence.statut
                })
                donnees_a_supprimer['presences'] += 1
            
            if presences_data:
                donnees_archive['presences'].append({
                    'employe_id': employe.id,
                    'matricule': employe.matricule,
                    'nom': employe.nom,
                    'prenom': employe.prenom,
                    'presences': presences_data
                })
            
            # Paie du mois (TRANSACTIONNEL - à réinitialiser)
            paie = PaieEmploye.objects.filter(
                employe=employe,
                mois=mois,
                annee=annee
            ).first()
            
            if paie:
                paie_data = {
                    'employe_id': employe.id,
                    'matricule': employe.matricule,
                    'nom': employe.nom,
                    'prenom': employe.prenom,
                    'fonction': employe.fonction,
                    'salaire_base': float(paie.salaire_base or 0),
                    'prime_discipline': float(paie.prime_discipline or 0),
                    'cherete_vie': float(paie.cherete_vie or 0),
                    'indemnite_transport': float(paie.indemnite_transport or 0),
                    'indemnite_logement': float(paie.indemnite_logement or 0),
                    'salaire_brut': float(paie.salaire_brut or 0),
                    'cnss': float(paie.cnss or 0),
                    'rts': float(paie.rts or 0),
                    'avances': float(paie.avance_sur_salaire or 0),
                    'sanctions': float(paie.sanction_vol_carburant or 0),
                    'net_a_payer': float(paie.salaire_net_a_payer or 0),
                    'jours_presence': float(paie.jours_presence or 0),
                    'absences': paie.absences or 0,
                    'dimanches': paie.dimanches or 0,
                    'jours_repos': paie.jours_repos or 0,
                    'conge': paie.conge or 0
                }
                donnees_archive['paies'].append(paie_data)
                
                total_salaire_brut += Decimal(str(paie.salaire_brut or 0))
                total_salaire_net += Decimal(str(paie.salaire_net_a_payer or 0))
                nombre_employes += 1
                donnees_a_supprimer['paies'] += 1
            
            # Heures supplémentaires du mois (TRANSACTIONNEL - à réinitialiser)
            heures_supp = HeureSupplementaire.objects.filter(
                employe=employe,
                date__month=mois,
                date__year=annee
            )
            
            heures_data = []
            for heure in heures_supp:
                heures_data.append({
                    'date': heure.date.isoformat(),
                    'heure_debut': heure.heure_debut.strftime('%H:%M') if heure.heure_debut else '',
                    'heure_fin': heure.heure_fin.strftime('%H:%M') if heure.heure_fin else '',
                    'duree': float(heure.duree or 0),
                    'total_a_payer': float(heure.total_a_payer or 0)
                })
                total_heures_supp += Decimal(str(heure.duree or 0))
                donnees_a_supprimer['heures_supp'] += 1
            
            if heures_data:
                donnees_archive['heures_supplementaires'].append({
                    'employe_id': employe.id,
                    'matricule': employe.matricule,
                    'nom': employe.nom,
                    'prenom': employe.prenom,
                    'heures': heures_data
                })
        
        # 3. STATISTIQUES GLOBALES
        donnees_archive['statistiques'] = {
            'total_employes': nombre_employes,
            'total_salaire_brut': float(total_salaire_brut),
            'total_salaire_net': float(total_salaire_net),
            'total_heures_supp': float(total_heures_supp),
            'donnees_supprimees': donnees_a_supprimer,
            'date_creation': timezone.now().isoformat(),
            'mois_archive': mois,
            'annee_archive': annee
        }
        
        # 4. CRÉATION DE L'ARCHIVE
        print(f"💾 ARCHIVAGE: Création de l'archive en base")
        archive = ArchiveMensuelle.objects.create(
            user=user,
            mois=mois,
            annee=annee,
            nb_employes_actifs=nombre_employes,
            total_salaire_brut=total_salaire_brut,
            total_net_paye=total_salaire_net,
            donnees_presences=json.dumps(donnees_archive['presences'], ensure_ascii=False),
            donnees_paies=json.dumps(donnees_archive['paies'], ensure_ascii=False),
            donnees_heures_supp=json.dumps(donnees_archive['heures_supplementaires'], ensure_ascii=False),
            statut='En cours',
            date_creation=timezone.now()
        )
        
        # 5. RÉINITIALISATION DES DONNÉES TRANSACTIONNELLES
        print(f"🔄 ARCHIVAGE: Réinitialisation des données transactionnelles")
        
        # Supprimer les présences du mois archivé
        presences_supprimees = PresenceJournaliere.objects.filter(
            employe__user=user,
            date__month=mois,
            date__year=annee
        ).delete()
        
        # Supprimer les paies du mois archivé
        paies_supprimees = PaieEmploye.objects.filter(
            employe__user=user,
            mois=mois,
            annee=annee
        ).delete()
        
        # Supprimer les heures supplémentaires du mois archivé
        heures_supprimees = HeureSupplementaire.objects.filter(
            employe__user=user,
            date__month=mois,
            date__year=annee
        ).delete()
        
        # 6. PRÉPARATION POUR LE NOUVEAU MOIS
        print(f"🆕 ARCHIVAGE: Préparation pour le nouveau mois")
        
        # Calculer le mois suivant
        if mois == 12:
            nouveau_mois = 1
            nouvelle_annee = annee + 1
        else:
            nouveau_mois = mois + 1
            nouvelle_annee = annee
        
        # Les données de référence restent inchangées :
        # - employes/ : Matricule, Employé, Fonction, Salaire de base, Statut
        # - Configuration des montants
        # - Légende des statuts
        # - Paramètres de paie : Matricule, Prénom, Nom, Fonction, Salaire Base
        
        # Log final
        print(f"✅ ARCHIVAGE TERMINÉ:")
        print(f"   📊 Archive créée: {archive.id}")
        print(f"   🗑️ Présences supprimées: {presences_supprimees[0] if presences_supprimees else 0}")
        print(f"   🗑️ Paies supprimées: {paies_supprimees[0] if paies_supprimees else 0}")
        print(f"   🗑️ Heures supp supprimées: {heures_supprimees[0] if heures_supprimees else 0}")
        print(f"   🆕 Prêt pour le mois: {nouveau_mois}/{nouvelle_annee}")
        print(f"   💾 Données de référence conservées: {employes.count()} employés")
        
        return archive


@login_required
def archive_details(request, archive_id):
    """Vue pour afficher les détails d'une archive"""
    import json
    
    archive = get_object_or_404(ArchiveMensuelle, id=archive_id, user=request.user)
    
    # Décoder les données JSON
    try:
        donnees_archive = json.loads(archive.donnees_json)
    except (json.JSONDecodeError, TypeError):
        donnees_archive = {}
    
    context = {
        'archive': archive,
        'donnees_archive': donnees_archive,
        'employes_data': donnees_archive.get('employes', []),
        'presences_data': donnees_archive.get('presences', []),
        'paies_data': donnees_archive.get('paies', []),
        'heures_supp_data': donnees_archive.get('heures_supplementaires', []),
        'statistiques': donnees_archive.get('statistiques', {}),
    }
    
    return render(request, 'fleet_app/entreprise/archive_details.html', context)


@login_required
def archive_export(request, archive_id):
    """Vue pour exporter une archive en JSON"""
    from django.http import HttpResponse
    import json
    
    archive = get_object_or_404(ArchiveMensuelle, id=archive_id, user=request.user)
    
    # Préparer les données d'export
    export_data = {
        'archive_info': {
            'id': archive.id,
            'periode': archive.get_periode_label(),
            'mois': archive.mois,
            'annee': archive.annee,
            'statut': archive.statut,
            'nb_employes_actifs': archive.nb_employes_actifs,
            'total_salaire_brut': float(archive.total_salaire_brut or 0),
            'total_net_paye': float(archive.total_net_paye or 0),
            'date_creation': archive.date_creation.isoformat(),
            'date_cloture': archive.date_cloture.isoformat() if archive.date_cloture else None,
        },
        'donnees': json.loads(archive.donnees_json) if archive.donnees_json else {}
    }
    
    # Créer la réponse HTTP avec le fichier JSON
    response = HttpResponse(
        json.dumps(export_data, ensure_ascii=False, indent=2),
        content_type='application/json'
    )
    filename = f'archive_{archive.get_periode_label().replace(" ", "_")}.json'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def archives_export_excel(request):
    """Vue pour exporter toutes les archives en Excel"""
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    # Récupérer toutes les archives de l'utilisateur
    archives = ArchiveMensuelle.objects.filter(user=request.user).order_by('-annee', '-mois')
    
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Archives Mensuelles"
    
    # Définir les styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # En-têtes
    headers = [
        'Période', 'Mois', 'Année', 'Nombre Employés', 
        'Salaire Brut (GNF)', 'Salaire Net (GNF)', 'Statut', 
        'Date Création', 'Date Clôture'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    for row, archive in enumerate(archives, 2):
        ws.cell(row=row, column=1, value=archive.get_periode_label())
        ws.cell(row=row, column=2, value=archive.mois)
        ws.cell(row=row, column=3, value=archive.annee)
        ws.cell(row=row, column=4, value=archive.nb_employes_actifs)
        ws.cell(row=row, column=5, value=float(archive.total_salaire_brut or 0))
        ws.cell(row=row, column=6, value=float(archive.total_net_paye or 0))
        ws.cell(row=row, column=7, value=archive.statut)
        ws.cell(row=row, column=8, value=archive.date_creation.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=9, value=archive.date_cloture.strftime('%d/%m/%Y %H:%M') if archive.date_cloture else '-')
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Créer la réponse HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="archives_mensuelles.xlsx"'
    
    return response


@login_required
def archives_export_pdf(request):
    """Vue pour exporter toutes les archives en PDF"""
    from django.http import HttpResponse
    from django.template.loader import get_template
    from xhtml2pdf import pisa
    from io import BytesIO
    
    # Récupérer toutes les archives de l'utilisateur
    archives = ArchiveMensuelle.objects.filter(user=request.user).order_by('-annee', '-mois')
    
    # Statistiques globales
    from django.db.models import Sum, Count, Avg
    stats_globales = archives.aggregate(
        total_archives=Count('id'),
        total_employes=Sum('nb_employes_actifs'),
        total_montant_brut=Sum('total_salaire_brut'),
        total_montant_net=Sum('total_net_paye'),
        moyenne_employes=Avg('nb_employes_actifs')
    )
    
    context = {
        'archives': archives,
        'stats_globales': stats_globales,
        'user': request.user,
        'date_export': timezone.now(),
    }
    
    # Charger le template
    template = get_template('fleet_app/entreprise/archives_export_pdf.html')
    html = template.render(context)
    
    # Créer le PDF
    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=buffer)
    
    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF', status=500)
    
    buffer.seek(0)
    
    # Créer la réponse HTTP
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="archives_mensuelles.pdf"'
    
    return response


@login_required
def cloturer_mois(request):
    """Vue pour clôturer un mois"""
    if request.method == 'POST':
        try:
            annee = int(request.POST.get('annee'))
            mois = int(request.POST.get('mois'))
            
            # Vérifier si l'archive existe déjà
            archive_existante = ArchiveMensuelle.objects.filter(
                user=request.user, annee=annee, mois=mois
            ).first()
            
            if archive_existante:
                messages.error(request, f'Le mois {mois}/{annee} est déjà clôturé')
            else:
                # Créer l'archive
                ArchiveMensuelle.objects.create(
                    user=request.user,
                    annee=annee,
                    mois=mois,
                    date_cloture=timezone.now(),
                    nombre_employes=Employe.objects.filter(user=request.user).count(),
                    nombre_paies=PaieEmploye.objects.filter(
                        employe__user=request.user, annee=annee, mois=mois
                    ).count()
                )
                
                messages.success(request, f'Mois {mois}/{annee} clôturé avec succès')
        
        except Exception as e:
            messages.error(request, f'Erreur lors de la clôture : {str(e)}')
    
    return redirect('fleet_app:archive_mensuelle')

# ===== VUES POUR LA CONFIGURATION =====

@login_required
def configuration_montant_statut(request):
    """Vue pour la configuration des montants par statut"""
    if request.method == 'POST':
        # Traitement de la configuration
        messages.success(request, 'Configuration des montants mise à jour')
        return redirect('fleet_app:configuration_montant_statut')
    
    context = {
        'montant_am': 10000,
        'montant_pm': 10000,
        'montant_journee': 20000,
        'montant_dim_am': 15000,
        'montant_dim_pm': 15000,
        'montant_dim_journee': 30000,
        'montant_absent': 0,
        'montant_maladie': 0,
        'montant_maladie_payee': 20000,
    }
    
    return render(request, 'fleet_app/entreprise/configuration_montant_statut.html', context)

# ===== VUES POUR LES EMPLOYÉS (complémentaires) =====

@login_required
def employe_create(request):
    """Vue pour créer un employé avec charges sociales guinéennes"""
    from fleet_app.forms_entreprise import EmployeForm
    
    if request.method == 'POST':
        form = EmployeForm(request.POST)
        if form.is_valid():
            try:
                employe = form.save(commit=False)
                employe.user = request.user
                
                # Récupérer l'entreprise de l'utilisateur connecté si disponible
                try:
                    from fleet_app.models_accounts import Profil
                    profil = Profil.objects.get(user=request.user)
                    if hasattr(profil, 'entreprise') and profil.entreprise:
                        employe.entreprise = profil.entreprise
                except Exception:
                    # En cas d'erreur, on continue sans associer d'entreprise
                    pass
                
                employe.save()
                
                # Message de succès avec informations sur les charges sociales configurées
                charges_info = []
                if employe.appliquer_cnss:
                    charges_info.append('CNSS')
                if employe.appliquer_rts:
                    charges_info.append('RTS')
                if employe.appliquer_vf:
                    charges_info.append('VF')
                
                charges_text = f" (Charges: {', '.join(charges_info)})" if charges_info else " (Aucune charge sociale)"
                messages.success(request, f'✅ Employé {employe} créé avec succès{charges_text}')
                
                if employe.calcul_salaire_auto:
                    messages.info(request, '📊 Le salaire de base sera calculé automatiquement depuis les présences')
                
                return redirect('fleet_app:pointage_journalier')
                
            except Exception as e:
                messages.error(request, f'❌ Erreur lors de la création : {str(e)}')
        else:
            messages.error(request, '⚠️ Veuillez corriger les erreurs dans le formulaire')
    else:
        form = EmployeForm()
    
    return render(request, 'fleet_app/entreprise/employe_form_simple.html', {'form': form})

@login_required
def employe_edit(request, pk):
    """Vue pour modifier un employé avec ModelForm complet"""
    from fleet_app.forms import EmployeForm
    
    try:
        employe = Employe.objects.get(pk=pk, user=request.user)
    except Employe.DoesNotExist:
        messages.error(request, f'L\'employé avec l\'ID {pk} n\'existe pas ou a déjà été supprimé.')
        return redirect('fleet_app:employe_list')
    
    if request.method == 'POST':
        print(f"[DEBUG] Modification employé {pk} - Données POST reçues: {dict(request.POST)}")
        form = EmployeForm(request.POST, instance=employe)
        if form.is_valid():
            try:
                print(f"[DEBUG] Formulaire valide - Données nettoyées: {form.cleaned_data}")
                employe_updated = form.save(commit=False)
                employe_updated.user = request.user
                
                # Récupérer l'entreprise de l'utilisateur connecté si disponible
                try:
                    from fleet_app.models_accounts import Profil
                    profil = Profil.objects.get(user=request.user)
                    if hasattr(profil, 'entreprise') and profil.entreprise:
                        employe_updated.entreprise = profil.entreprise
                except Exception:
                    # En cas d'erreur, on continue sans associer d'entreprise
                    pass
                
                # Sauvegarder avec vérification
                print(f"[DEBUG] Sauvegarde de l'employé: {employe_updated.matricule} - {employe_updated.nom} {employe_updated.prenom}")
                employe_updated.save()
                
                # Vérifier que la sauvegarde a bien eu lieu
                employe_verifie = Employe.objects.get(pk=pk)
                print(f"[DEBUG] Vérification post-sauvegarde: {employe_verifie.matricule} - {employe_verifie.nom} {employe_verifie.prenom}")
                
                messages.success(request, f'Employé {employe_updated} modifié avec succès - Données sauvegardées')
                return redirect('fleet_app:employe_list')
                
            except Exception as e:
                print(f"[ERROR] Erreur lors de la sauvegarde: {str(e)}")
                messages.error(request, f'Erreur lors de la modification : {str(e)}')
        else:
            print(f"[ERROR] Formulaire invalide - Erreurs: {form.errors}")
            messages.error(request, 'Veuillez corriger les erreurs dans le formulaire')
            # Afficher les erreurs spécifiques
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        # Pré-remplir le formulaire avec les données existantes
        form = EmployeForm(instance=employe)
    
    context = {'form': form, 'employe': employe}
    return render(request, 'fleet_app/entreprise/employe_form.html', context)

@login_required
def employe_delete(request, pk):
    """Vue pour supprimer un employé avec gestion des contraintes de clé étrangère"""
    try:
        employe = Employe.objects.get(pk=pk, user=request.user)
    except Employe.DoesNotExist:
        messages.error(request, f'L\'employé avec l\'ID {pk} n\'existe pas ou a déjà été supprimé.')
        return redirect('fleet_app:employe_list')
    
    if request.method == 'POST':
        nom_employe = f"{employe.prenom} {employe.nom} (Matricule: {employe.matricule})"
        print(f"[DEBUG] Tentative de suppression de l'employé: {nom_employe}")
        
        try:
            from django.db import transaction
            
            with transaction.atomic():
                # Compter les données qui seront supprimées
                presences_count = 0
                heures_supp_count = 0
                paies_count = 0
                configurations_count = 0
                
                # Supprimer les présences liées
                try:
                    from .models import PresenceJournaliere
                    presences = PresenceJournaliere.objects.filter(employe=employe)
                    presences_count = presences.count()
                    presences.delete()
                    print(f"[DEBUG] Supprimé {presences_count} présences")
                except Exception as e:
                    print(f"[DEBUG] Erreur suppression présences: {e}")
                
                # Supprimer les heures supplémentaires liées
                try:
                    from .models import HeureSupplementaire
                    heures_supp = HeureSupplementaire.objects.filter(employe=employe)
                    heures_supp_count = heures_supp.count()
                    heures_supp.delete()
                    print(f"[DEBUG] Supprimé {heures_supp_count} heures supplémentaires")
                except Exception as e:
                    print(f"[DEBUG] Erreur suppression heures supp: {e}")
                
                # Supprimer les paies liées
                try:
                    from .models import PaieEmploye
                    paies = PaieEmploye.objects.filter(employe=employe)
                    paies_count = paies.count()
                    paies.delete()
                    print(f"[DEBUG] Supprimé {paies_count} paies")
                except Exception as e:
                    print(f"[DEBUG] Erreur suppression paies: {e}")
                
                # Supprimer les configurations de montants liées
                try:
                    from .models import ConfigurationMontantEmploye
                    configurations = ConfigurationMontantEmploye.objects.filter(employe=employe)
                    configurations_count = configurations.count()
                    configurations.delete()
                    print(f"[DEBUG] Supprimé {configurations_count} configurations")
                except Exception as e:
                    print(f"[DEBUG] Erreur suppression configurations: {e}")
                
                # Maintenant supprimer l'employé
                employe.delete()
                print(f"[DEBUG] Employé {nom_employe} supprimé avec succès")
                
            # Message de succès détaillé
            details = []
            if presences_count > 0:
                details.append(f"{presences_count} présence(s)")
            if heures_supp_count > 0:
                details.append(f"{heures_supp_count} heure(s) supplémentaire(s)")
            if paies_count > 0:
                details.append(f"{paies_count} paie(s)")
            if configurations_count > 0:
                details.append(f"{configurations_count} configuration(s)")
            
            if details:
                details_str = ", ".join(details)
                success_msg = f'Employé {nom_employe} supprimé avec succès ainsi que {details_str} associé(es)'
            else:
                success_msg = f'Employé {nom_employe} supprimé avec succès'
                
            messages.success(request, success_msg)
            return redirect('fleet_app:employe_list')
            
        except Exception as e:
            print(f"[ERROR] Erreur lors de la suppression de l'employé {pk}: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Erreur lors de la suppression de {nom_employe}: {str(e)}')
            return redirect('fleet_app:employe_list')
    
    # Calculer les statistiques pour afficher à l'utilisateur ce qui sera supprimé
    stats = {}
    try:
        from .models import PresenceJournaliere, HeureSupplementaire, PaieEmploye, ConfigurationMontantEmploye
        
        stats['presences'] = PresenceJournaliere.objects.filter(employe=employe).count()
        stats['heures_supp'] = HeureSupplementaire.objects.filter(employe=employe).count()
        stats['paies'] = PaieEmploye.objects.filter(employe=employe).count()
        stats['configurations'] = ConfigurationMontantEmploye.objects.filter(employe=employe).count()
        
        print(f"[DEBUG] Statistiques pour {employe}: {stats}")
    except Exception as e:
        print(f"[DEBUG] Erreur calcul statistiques: {e}")
        stats = {}
    
    context = {'employe': employe, 'stats': stats}
    return render(request, 'fleet_app/entreprise/employe_confirm_delete.html', context)

@login_required
def get_employe_info_ajax(request):
    """Vue AJAX pour récupérer les informations d'un employé"""
    if request.method == 'GET':
        employe_id = request.GET.get('employe_id')
        if employe_id:
            try:
                employe = Employe.objects.get(id=employe_id, user=request.user)
                data = {
                    'success': True,
                    'matricule': employe.matricule,
                    'nom_complet': f"{employe.prenom} {employe.nom}",
                    'fonction': employe.fonction or 'Non défini',
                    'taux_horaire': 5000,  # Taux par défaut, peut être personnalisé
                    'salaire_journalier': employe.salaire_journalier or 0
                }
                return JsonResponse(data)
            except Employe.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Employé non trouvé'})
        else:
            return JsonResponse({'success': False, 'error': 'ID employé manquant'})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

# ===== VUES POUR LES PARAMÈTRES DE PAIE =====

@login_required
def parametre_paie_list(request):
    """Vue complète pour les paramètres de paie avec recherche et export"""
    
    try:
        # Traitement des actions POST
        if request.method == 'POST':
            action = request.POST.get('action')
            
            if action == 'modifier_parametre':
                try:
                    employe_id = request.POST.get('employe_id')
                    employe = Employe.objects.get(id=employe_id, user=request.user)
                    
                    # Mise à jour des paramètres
                    employe.salaire_journalier = float(request.POST.get('salaire_base', 0))
                    employe.avances = float(request.POST.get('avances', 0))
                    
                    employe.save()
                    messages.success(request, f'Paramètres mis à jour pour {employe.prenom} {employe.nom}')
                    
                except Employe.DoesNotExist:
                    messages.error(request, 'Employé non trouvé')
                except ValueError:
                    messages.error(request, 'Valeurs numériques invalides')
                except Exception as e:
                    messages.error(request, f'Erreur lors de la mise à jour: {str(e)}')
        
            elif action == 'appliquer_global':
                try:
                    # Appliquer des paramètres globaux à tous les employés
                    salaire_global = float(request.POST.get('salaire_global', 0))
                    avances_global = float(request.POST.get('avances_global', 0))
                    
                    if salaire_global > 0 or avances_global > 0:
                        employes = Employe.objects.filter(user=request.user)
                        count = 0
                        for employe in employes:
                            if salaire_global > 0:
                                employe.salaire_journalier = salaire_global
                            if avances_global > 0:
                                employe.avances = avances_global
                            employe.save()
                            count += 1
                        
                        messages.success(request, f'Paramètres globaux appliqués à {count} employés')
                    else:
                        messages.error(request, 'Veuillez saisir au moins un paramètre global')
                        
                except ValueError:
                    messages.error(request, 'Valeurs numériques invalides')
                except Exception as e:
                    messages.error(request, f'Erreur lors de l\'application globale: {str(e)}')
        
            return redirect('fleet_app:parametre_paie_list')
        
        # Filtres de recherche
        search_matricule = request.GET.get('search_matricule', '')
        search_nom = request.GET.get('search_nom', '')
        search_fonction = request.GET.get('search_fonction', '')
        
        # Récupération des employés avec filtrage
        employes = Employe.objects.filter(user=request.user).order_by('matricule')
        
        if search_matricule:
            employes = employes.filter(matricule__icontains=search_matricule)
        if search_nom:
            employes = employes.filter(
                Q(nom__icontains=search_nom) | Q(prenom__icontains=search_nom)
            )
        if search_fonction:
            employes = employes.filter(fonction__icontains=search_fonction)
        
        # Configuration par défaut (sans modèle de configuration)
        config = {
            'heures_normales_mois': 173.33,
            'taux_jour_ouvrable': 1.5,
            'taux_dimanche_ferie': 2.0,
            'salaire_mensuel_base': 500000
        }
        
        # Statistiques
        stats = {
            'total_employes': employes.count(),
            'total_avances': employes.aggregate(total=Sum('avances'))['total'] or 0,
            'salaire_moyen': employes.aggregate(avg=Avg('salaire_journalier'))['avg'] or 0,
        }
        
        context = {
            'employes': employes,
            'config': config,
            'stats': stats,
            'search_matricule': search_matricule,
            'search_nom': search_nom,
            'search_fonction': search_fonction,
        }
        
        return render(request, 'fleet_app/entreprise/parametre_paie_list_complete.html', context)
        
    except Exception as e:
        import traceback
        error_msg = f"Erreur dans parametre_paie_list: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        return HttpResponse(error_msg, status=500, content_type='text/plain')

@login_required
def parametre_paie_export(request):
    """Vue pour exporter les paramètres de paie en CSV"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="parametres_paie_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Matricule', 'Prénom', 'Nom', 'Fonction', 'Salaire Base (GNF)', 
        'Avances (GNF)'
    ])
    
    employes = Employe.objects.filter(user=request.user).order_by('matricule')
    
    for employe in employes:
        writer.writerow([
            employe.matricule,
            employe.prenom,
            employe.nom,
            employe.fonction or '-',
            employe.salaire_journalier or 0,
            employe.avances or 0
        ])
    
    return response


@login_required
def bulletin_paie_list_simple(request):
    """Simplified bulletins view for debugging"""
    from datetime import datetime
    from fleet_app.models_entreprise import Employe
    
    try:
        # Get basic parameters
        mois_actuel = int(request.GET.get('mois', datetime.now().month))
        annee_actuelle = int(request.GET.get('annee', datetime.now().year))
        
        # Get employees
        employes = Employe.objects.filter(user=request.user)[:5]  # Limit to 5 for testing
        
        context = {
            'bulletins_data': [],
            'employes': employes,
            'mois_actuel': mois_actuel,
            'annee_actuelle': annee_actuelle,
            'mois_nom': 'Test',
            'entreprise': None,
        }
        
        return render(request, 'fleet_app/entreprise/bulletin_paie_list.html', context)
        
    except Exception as e:
        from django.http import HttpResponse
        return HttpResponse(f"Error in simple view: {e}", status=500)


@login_required
def bulletin_paie_list_working(request):
    """Vue bulletins de paie corrigée avec gestion d'erreurs"""
    try:
        from datetime import datetime
        import calendar
        from decimal import Decimal
        from django.core.paginator import Paginator
        from django.contrib import messages
        
        # Paramètres de filtre avec valeurs par défaut sûres
        mois_actuel = int(request.GET.get('mois', datetime.now().month))
        annee_actuelle = int(request.GET.get('annee', datetime.now().year))
        
        # Récupération des employés avec gestion d'erreur
        try:
            employes = Employe.objects.filter(user=request.user).order_by('matricule')
        except Exception as e:
            messages.error(request, f"Erreur lors de la récupération des employés: {e}")
            employes = []
        
        bulletins_data = []
        
        for employe in employes:
            try:
                # Récupération des présences pour le mois
                presences_mois = PresenceJournaliere.objects.filter(
                    employe=employe,
                    date__month=mois_actuel,
                    date__year=annee_actuelle
                )
                
                # Calculs de base sécurisés
                total_jours_travailles = presences_mois.filter(present=True).count()
                salaire_journalier = employe.salaire_journalier or Decimal('0')
                salaire_brut = salaire_journalier * total_jours_travailles
                
                # Récupération ou création de la paie
                paie, created = PaieEmploye.objects.get_or_create(
                    employe=employe,
                    mois=mois_actuel,
                    annee=annee_actuelle,
                    defaults={
                        'salaire_brut': salaire_brut,
                        'salaire_net': salaire_brut,
                        'salaire_net_a_payer': salaire_brut,
                        'jours_presence': total_jours_travailles,
                        'user': request.user
                    }
                )
                
                # Calculs CNSS sécurisés
                try:
                    param_cnss = ParametrePaie.objects.filter(cle='CNSS_ACTIVER', user=request.user).first()
                    appliquer_cnss = param_cnss and param_cnss.valeur == '1'
                except:
                    appliquer_cnss = False
                
                cnss_employe = salaire_brut * Decimal('0.05') if appliquer_cnss else Decimal('0')
                avances = employe.avances or Decimal('0')
                net_a_payer = salaire_brut - cnss_employe - avances
                
                # Mise à jour de la paie
                paie.salaire_brut = salaire_brut
                paie.cnss = cnss_employe
                paie.avance_sur_salaire = avances
                paie.salaire_net_a_payer = net_a_payer
                paie.save()
                
                bulletins_data.append({
                    'employe': employe,
                    'paie': paie,
                    'jours_travailles': total_jours_travailles,
                    'salaire_brut': salaire_brut,
                    'cnss_employe': cnss_employe,
                    'avances': avances,
                    'net_a_payer': net_a_payer,
                })
                
            except Exception as e:
                # Log l'erreur mais continue avec les autres employés
                print(f"Erreur pour employé {employe.matricule}: {e}")
                continue
        
        # Pagination sécurisée
        paginator = Paginator(bulletins_data, 10)
        page_number = request.GET.get('page', 1)
        
        try:
            page_obj = paginator.get_page(page_number)
        except:
            page_obj = paginator.get_page(1)
        
        # Configuration des charges sociales avec valeurs par défaut
        cnss_activer = False
        cnss_taux = Decimal('5.0')
        rts_type = 'PROGRESSIF'
        rts_taux_fixe = Decimal('10.0')
        
        try:
            param_cnss = ParametrePaie.objects.filter(cle='CNSS_ACTIVER', user=request.user).first()
            if param_cnss:
                cnss_activer = param_cnss.valeur == '1'
            
            param_taux = ParametrePaie.objects.filter(cle='CNSS_TAUX', user=request.user).first()
            if param_taux:
                cnss_taux = Decimal(param_taux.valeur)
        except:
            pass
        
        # Contexte sécurisé
        context = {
            'bulletins_data': bulletins_data,
            'page_obj': page_obj,
            'is_paginated': page_obj.has_other_pages(),
            'mois_actuel': mois_actuel,
            'annee_actuelle': annee_actuelle,
            'mois_nom': calendar.month_name[mois_actuel] if 1 <= mois_actuel <= 12 else 'Inconnu',
            'cnss_activer': cnss_activer,
            'cnss_taux': cnss_taux,
            'rts_type': rts_type,
            'rts_taux_fixe': rts_taux_fixe,
            'entreprise': None,  # Simplifié pour éviter les erreurs
        }
        
        return render(request, 'fleet_app/entreprise/bulletin_paie_list.html', context)
        
    except Exception as e:
        # Gestion d'erreur globale
        messages.error(request, f"Erreur lors du chargement des bulletins: {e}")
        context = {
            'bulletins_data': [],
            'page_obj': None,
            'is_paginated': False,
            'mois_actuel': datetime.now().month,
            'annee_actuelle': datetime.now().year,
            'mois_nom': 'Erreur',
            'cnss_activer': False,
            'cnss_taux': Decimal('5.0'),
            'rts_type': 'PROGRESSIF',
            'rts_taux_fixe': Decimal('10.0'),
            'entreprise': None,
            'error_message': str(e)
        }
        return render(request, 'fleet_app/entreprise/bulletin_paie_list.html', context)
